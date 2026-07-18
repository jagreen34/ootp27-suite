"""
draft_export.py — flat spreadsheet export of the draft board.

One row per prospect. Every raw rating (CON/POW/GAP/EYE + pitch grades) PLUS the
derived engine fields (career/disc/growth_bet/tier/parkfit/flags...) flattened into
one wide table, so the user can sort/filter/pivot in Excel and do their own work.

Reuses build_board's output — does NOT recompute value. New export path only;
touches nothing in the existing board. Follows suite conventions: _s accessor,
fail-loud on missing pool, glyphs already baked into the string fields.

Wire-up (in draft.py, inside render_draft or a button):
    from draft_export import board_to_dataframe, write_board_xlsx
    df = board_to_dataframe(rows, pool_df)           # rows = build_board(...)
    path = write_board_xlsx(df, league)              # -> .xlsx path
    st.download_button("Export board (.xlsx)", data=open(path,'rb'), file_name=...)
"""
from __future__ import annotations
import pandas as pd
from acquisitions import _s

# ── Raw-rating columns to surface, in a sensible reading order ────────────────
# Batter ratings (current + potential + the vL/vR splits kept optional/compact)
BAT_RATING_COLS = [
    ('CON', 'CON'), ('POW', 'POW'), ('GAP', 'GAP'), ('EYE', 'EYE'),
    ('BABIP', 'BABIP'), ('SPE', 'SPE'), ('STE', 'STE'),
    ('CON P', 'CON_POT'), ('POW P', 'POW_POT'), ('GAP P', 'GAP_POT'), ('EYE P', 'EYE_POT'),
]
# Pitcher ratings (current + potential) — pitch GRADES handled separately by
# ALL_PITCH_COLS (all 12 types), so only the non-pitch ratings live here.
PIT_RATING_COLS = [
    ('STU', 'STU'), ('MOV', 'MOV'), ('HRA', 'HRA'), ('PIT_CON', 'PCON'),
    ('PBABIP', 'PBABIP'), ('STM', 'STM'), ('velo_mid', 'VELO'),
    ('PIT_CON_P', 'PCON_POT'),
]
# Pitcher counting/box columns already in the pool, handy for slicing
PIT_STAT_COLS = [
    ('PIT_K_PCT', 'K%'), ('PIT_BB_PCT', 'BB%'), ('PIT_GF', 'GB/FB'),
]

# ALL pitch-grade columns — using the EXACT processed names from acquisitions
# PLAYER_RENAMES (verified against the real draft-pool export). prep_data renames
# the short CSV names (FB/CB/SP/FO/SC/KC/KN...) → PIT_* BEFORE build_board runs, so
# by the time this export sees the row, columns are in PIT_* form. F2 RETRAIN:
# the value model now scores ALL 12 grades (flat, free coefs) — nothing is
# model-blind. All 12 surfaced as DISPLAY columns and scored.
ALL_PITCH_COLS = [
    ('PIT_FB_GR', 'FB'),    # fastball  (modeled)
    ('PIT_SI',    'SI'),    # sinker    (modeled)
    ('PIT_CT',    'CT'),    # cutter
    ('PIT_CB',    'CB'),    # curveball  (modeled)
    ('PIT_SL',    'SL'),    # slider    (modeled)
    ('PIT_CH',    'CH'),    # changeup  (modeled)
    ('PIT_SP',    'SP'),    # splitter   (modeled)
    ('PIT_FO',    'FO'),    # forkball
    ('PIT_CC',    'CC'),    # circle change
    ('PIT_SC',    'SC'),    # screwball
    ('PIT_KC',    'KC'),    # knuckle curve
    ('PIT_KN',    'KN'),    # knuckleball
]
# F2 RETRAIN: value model now scores ALL 12 pitch grades (flat, free coefs).
# OffModelBest flag RETIRED — no pitch is model-blind anymore.


def _pitch_col(row, col):
    """Return the rating for a pitch column (single exact name)."""
    if col in row and not pd.isna(row.get(col)):
        return _rating(row, col)
    return ''


def _rating(row, col):
    """Raw rating passthrough. Blank (not 0) when the column is absent, so the
    sheet shows an empty cell rather than a fake zero the user might sort on."""
    if col not in row or pd.isna(row.get(col)):
        return ''
    v = row.get(col)
    # ratings are ints on the 20-80 (or 1-100) scale; keep them as ints
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return _s(v)


# ── A31 age-budget cap (Option A: applied in DISPLAYED grade-points) ───────────
# What we found this session (A31): development is an AGE BUDGET with a cliff at
# 24-25, NOT the smooth linear delivery-fraction the old GrowthBet uses. A young
# player realistically climbs ~2 grades (21-23), then it craters. Effective
# ceiling = min(shown_potential, current + age_cap). Capped in displayed grade-
# points (eye-verifiable; most accurate at the top of the scale where decisions
# live). Pitcher cliff-age is confirmed only at 19 (A31 scope note) — arms use the
# same table but hold it a touch looser mentally.
#
# NOTE this is DELIBERATELY different from GrowthBet (disc − career), which is the
# old linear delivery-fraction model. This column is the age-budget/cliff finding.
AGE_CAP_GRADES = {
    # age : max displayed-grade points reachable above current
    #  ≤23 → +10 (2 grades) | 24 → +5 | 25 → +3 | 26+ → 0
}
def age_cap(age: float) -> int:
    a = _s(age, 99)
    if a <= 23:  return 10
    if a == 24:  return 5
    if a == 25:  return 3
    return 0

def eff_ceiling(cur, pot, age) -> str:
    """min(shown potential, current + age cap). Blank if inputs missing."""
    if cur in ('', None) or pd.isna(cur):
        return ''
    c = _s(cur, None)
    if c is None:
        return ''
    cap = age_cap(age)
    reachable = c + cap
    # if no potential shown, ceiling is just current + cap
    p = _s(pot, None) if pot not in ('', None) and not pd.isna(pot) else None
    ceil = min(p, reachable) if p is not None else reachable
    # never below current (a potential < current shouldn't drag the ceiling down)
    ceil = max(ceil, c)
    return int(round(ceil))


# developing ratings to cap: (current_col, potential_col, output_label)
BAT_DEV_RATINGS = [
    ('CON', 'CON P', 'CON_EC'), ('POW', 'POW P', 'POW_EC'),
    ('GAP', 'GAP P', 'GAP_EC'), ('EYE', 'EYE P', 'EYE_EC'),
]
PIT_DEV_RATINGS = [
    ('MOV', 'MOV P', 'MOV_EC'), ('PIT_CON', 'PIT_CON_P', 'PCON_EC'),
    ('HRA', 'HRA P', 'HRA_EC'),
    ('PIT_FB_GR', 'PIT_FB_GR_P', 'FB_EC'), ('PIT_CH', 'PIT_CH_P', 'CH_EC'),
    ('PIT_SI', 'PIT_SI_P', 'SI_EC'), ('PIT_SL', 'PIT_SL_P', 'SL_EC'),
    ('PIT_CB', 'PIT_CB_P', 'CB_EC'), ('PIT_SP', 'PIT_SP_P', 'SP_EC'),
]


def board_to_dataframe(rows: list[dict], pool_df: pd.DataFrame) -> pd.DataFrame:
    """Flatten build_board rows + raw pool ratings into one wide export table.

    `rows` is the list of dicts returned by draft.build_board (already sorted by
    career WAR / BPA). `pool_df` is the same pool build_board scored — used to pull
    the raw ratings out of each prospect's underlying row.
    """
    if not rows:
        raise ValueError("draft_export: empty board — nothing to export "
                         "(is the draft pool loaded?)")

    out = []
    for x in rows:
        r = x['row']                      # underlying pool Series
        is_pit = x['is_pit']

        rec = {
            # ── identity ──────────────────────────────────────────────────
            'Name':     x['name'],
            'Pos':      x['pos'],
            'BestFit':  x.get('bestfit') or '',
            'Age':      x['age'],
            'B':        x.get('bt') or '',
            'T':        x.get('throws') or '',
            'Org':      str(r.get('ORG', '')),
            # ── engine value (reused, not recomputed) ─────────────────────
            'Tier':     x['tier'],
            'CareerWAR': x['career'],
            'DiscWAR':  x['disc'],
            'GrowthBet': x['growth_bet'],   # disc − career = credited upside
            'Window':   x['window'],
            'TradeVal': x['tv'],
            'Need':     'Y' if x['need'] else '',
            # ── side lenses (additive; never reordered the BPA rank) ──────
            'Glove':    x.get('glove') if x.get('glove') is not None else '',
            'ParkFit':  x.get('parkfit') if x.get('parkfit') is not None else '',
            'Edge':     x.get('edge') if x.get('edge') is not None else '',
            'Conf':     x.get('conf') if x.get('conf') is not None else '',
            'TotalVal': x.get('tval') if x.get('tval') is not None else '',
            # ── flags / gates (plain text so the user can filter on them) ──
            'Flags':    x.get('flags') or '',
            'Skip':     x.get('skip') or '',
            'RPceiling': 'Y' if x.get('rp_ceiling') else '',
            # ── pitcher structural (A34 reads, already in the pool) ───────
            'NPitch':   x.get('npitch') if is_pit else '',
            'TopPitch': x.get('top_pitch') if is_pit else '',
            'Def':      x.get('def_sum') if not is_pit else '',
        }

        # ── raw ratings, by type ──────────────────────────────────────────
        if is_pit:
            for col, label in PIT_RATING_COLS + PIT_STAT_COLS:
                rec[label] = _rating(r, col)
            # A31 effective ceiling per developing rating (min(pot, cur+age_cap))
            for cur_c, pot_c, label in PIT_DEV_RATINGS:
                rec[label] = eff_ceiling(r.get(cur_c), r.get(pot_c), x['age'])
            # ALL 12 pitch grades as display columns (F2 RETRAIN scores all 12)
            for col, label in ALL_PITCH_COLS:
                rec[label] = _pitch_col(r, col)
        else:
            for col, label in BAT_RATING_COLS:
                rec[label] = _rating(r, col)
            for cur_c, pot_c, label in BAT_DEV_RATINGS:
                rec[label] = eff_ceiling(r.get(cur_c), r.get(pot_c), x['age'])

        out.append(rec)

    df = pd.DataFrame(out)
    return df


def write_board_xlsx(df: pd.DataFrame, league=None, path: str | None = None) -> str:
    """Write the flat board to a filter-ready .xlsx (bulk data, no formulas).

    Freezes the header + Name column, turns on autofilter, sizes columns. No
    color fills (user is colorblind; glyphs already carry meaning in-string).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    if path is None:
        tag = ''
        try:
            tag = f"_{league.season}" if league is not None and getattr(league, 'season', None) else ''
        except Exception:
            tag = ''
        path = f"/tmp/draft_board{tag}.xlsx"

    # bulk write via pandas, then style with openpyxl
    df.to_excel(path, index=False, sheet_name='Draft Board')

    wb = openpyxl.load_workbook(path)
    ws = wb['Draft Board']

    # header: bold, wrapped, frozen; freeze Name column too (B2)
    for cell in ws[1]:
        cell.font = Font(name='Arial', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions

    # body font + column widths (sized to content, capped)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 6), 22)

    wb.save(path)
    return path
